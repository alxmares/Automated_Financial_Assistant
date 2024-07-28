import easyocr
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
import re
import datetime

class Reader():
    def __init__(self):
        self.reader = easyocr.Reader(['es']) # Crear lector
        self.path = ""
        self.text = ""
        self.relevant_text = ""
        self.date = None
        
        
    def get_text(self, path):
        if path == self.path:
            return self.text
        
        self.path = path
        # Leer la imagen y extraer el texto
        results = self.reader.readtext(path)

        # Concatenar los resultados en una sola cadena
        self.text = [result[1] for result in results]

        return self.text
    
    def get_relevant_text(self):
        init = 0
        end = 0
        total_value = 0.0
        TOTAL_OBTAINED = False
        # Expresión regular para detectar fechas en formato xx/xx/xxxx
        date_pattern = re.compile(r'\b\d{2}/\d{2}/\d{4}\b')
        
        for i, w in enumerate(self.text):
            if w == "CANT" or w=="CANT .":
                init = i
            if w == "**" and TOTAL_OBTAINED == False:
                for j in range(i + 1, len(self.text)):
                    if self.text[j].replace(' ', '').replace('.', '').replace(',', '').isdigit():
                        total_value = float(self.text[j].replace(' ', '').replace(',', ''))
                        end = j + 1
                        break
                if end != 0:
                    TOTAL_OBTAINED = True
            
            if date_pattern.match(w):  # Verificar si es una fecha
                print(f"Fecha encontrada: {w}")
                self.date = w
            
        print("Precio Total: ", total_value)   
        
        self.relevant_text = self.text[init:end]
        return self.relevant_text 
    
    def create_table(self):
        processed_data = []
        current_row = []
        articulo = []
        cantidad_encontrada = False
        for word in self.relevant_text:
            cleaned_word = word.replace(' ', '')  # Eliminar los espacios
            #print(cleaned_word, end='')
            
            if cleaned_word.replace('.', '', 1).isdigit() and float(cleaned_word) <= 100 and not cantidad_encontrada:  # Verificar si es cantidad
                if current_row:
                    current_row.append(" ".join(articulo))  # Unir las palabras del artículo
                articulo = []
                    
                current_row.append(cleaned_word)
                cantidad_encontrada = True
                #print(" Es cantidad")
                
            elif cleaned_word.replace('.', '', 1).isdigit() and len(current_row) == 1:  # Verificar si es precio
                #print(articulo)
                current_row.append(" ".join(articulo))  # Añadir artículo antes del primer precio
                articulo = []
                current_row.append(cleaned_word)
                #print(" Es precio")
                
            elif cleaned_word.replace('.', '', 1).isdigit() and len(current_row) == 3:  # Verificar si es total
                current_row.append(cleaned_word)
                #print(" Es Total")
            
            elif cleaned_word.replace('.', '', 1).isdigit() and len(current_row) > 3:  # Ajustar precio y total si hay tres números
                current_row[2] = current_row[3]  # Mover el primer precio al segundo
                current_row[3] = cleaned_word  # Guardar el nuevo total
                #print("Recorrer")    
            elif cleaned_word in ['A', 'B', 'K']:  # Verificar si es A, B o K
                current_row.append(cleaned_word)
                if len(current_row) == 5:
                    processed_data.append(current_row)
                current_row = []
                cantidad_encontrada = False
                #print(" Es Terminal")
                
            else:
                articulo.append(word)  # Acumular las palabras del artículo
                #print(" Es Palabra")
                
            #print(current_row)
        
        self.table = pd.DataFrame(processed_data, columns=["Cantidad", "Artículo", "Precio", "Total", "Tipo"])
            
        #print(self.table.head().to_markdown())
            
        return self.table
    
    def add_to_excel(self, file, sheet, today_date=False, date=None):
        if today_date:
            self.date = datetime.today().strftime('%d/%m/%Y')
        elif date:
            self.date = date

        try:
            # Agregar la columna de fecha
            self.table["Fecha"] = pd.to_datetime(self.date, format="%d/%m/%Y").date()
        except Exception as e:
            print(f"La fecha no es válida o la tabla ya contiene una fecha: {e}")
            print(self.table.head().to_markdown())
            return

        self.table["Cantidad"] = self.table["Cantidad"].astype(float)
        self.table["Precio"] = self.table["Precio"].astype(float)
        self.table["Total"] = self.table["Total"].astype(float)

        # Reordenar las columnas para que FECHA sea la primera
        self.table = self.table[["Fecha", "Cantidad", "Artículo", "Precio", "Total"]]
        
        book = load_workbook(file)
        if sheet in book.sheetnames:
            # Leer la hoja existente en un DataFrame
            df_existente = pd.read_excel(file, sheet_name=sheet)
            
            # Combinar los DataFrames
            df_combinado = pd.concat([df_existente, self.table], ignore_index=True)
        else:
            # Si la hoja no existe, simplemente usa el nuevo DataFrame
            df_combinado = self.table

        # Escribir el DataFrame combinado a la hoja específica, comenzando desde (0,0)
        with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_combinado.to_excel(writer, sheet_name=sheet, index=False)

        print("Datos agregados exitosamente.")
        print(self.table.head().to_markdown())
